from openpyxl import Workbook, load_workbook

import pandas as pd
import csv




wb = load_workbook('employeedata.xlsx')

ws = wb.active


range = ws["B2":"B30"]

def original_xlsx():
  for cell in range:
    for x in cell:
      print(x.value)
original_xlsx()
print("**************************************** THE UPDATED XLSX FILE IS BELOW********************************************")

def modified_xlsx():
  for cell in range:
    for x in cell:
      text = x.value
      sufix = text.replace("helpinghands.cm","handsinhands.org")
      x.value = sufix
      print(sufix)
      wb.save('modifiedemployeedata.xlsx')
      

modified_xlsx()

print("*********************************************CSV*******************************************************************")



df = pd.read_csv("employeedata.csv")

print(df)

print("***********************************UPDATED CSV FILE**********************************************")


df['E-mail'] = df['E-mail'].str.replace('helpinghands.cm', 'handsinhands.org')
print(df)

df.to_csv("modifiedEmployee.csv", index=False)